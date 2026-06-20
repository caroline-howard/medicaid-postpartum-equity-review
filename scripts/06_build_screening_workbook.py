from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import pandas as pd

from common import DATA, ensure_dirs, read_csv_if_exists, require_columns, write_csv


INPUT_CSV = DATA / "manual" / "screening_decisions.csv"
OUTPUT_XLSX = DATA / "manual" / "screening_decisions.xlsx"
SCORED_CSV = DATA / "processed" / "scored_records.csv"
DEDUPLICATED_CSV = DATA / "processed" / "deduplicated_records.csv"
FULL_TEXT_CSV = DATA / "manual" / "full_text_review.csv"
EXPECTED_RECORD_COUNT = 211

DECISION_OPTIONS = ["include_for_full_text", "maybe", "exclude"]
EXCLUSION_REASON_OPTIONS = [
    "wrong_population",
    "wrong_policy_or_intervention",
    "not_medicaid_or_chip",
    "not_postpartum",
    "not_us_based",
    "no_relevant_outcome",
    "background_only",
    "opinion_without_data_or_policy_detail",
    "duplicate",
    "other",
]

SCREENING_COLUMNS = [
    "record_id",
    "title",
    "abstract",
    "year",
    "journal",
    "doi",
    "pmid",
    "url",
    "source_database",
    "relevance_score",
    "automation_suggestion",
    "human_title_abstract_decision",
    "human_title_abstract_exclusion_reason",
    "full_text_needed",
    "notes",
]

PRESERVE_SCREENING_COLUMNS = [
    "human_title_abstract_decision",
    "human_title_abstract_exclusion_reason",
    "full_text_needed",
    "notes",
]

FULL_TEXT_COLUMNS = [
    "record_id",
    "citation",
    "full_text_retrieved",
    "retrieval_notes",
    "human_full_text_decision",
    "full_text_exclusion_reason",
    "study_or_report_type",
    "include_in_evidence_table",
    "notes",
]

COLUMN_WIDTHS = {
    "record_id": 16,
    "title": 64,
    "abstract": 100,
    "year": 10,
    "journal": 32,
    "doi": 28,
    "pmid": 14,
    "url": 42,
    "source_database": 18,
    "relevance_score": 16,
    "automation_suggestion": 24,
    "human_title_abstract_decision": 32,
    "human_title_abstract_exclusion_reason": 42,
    "full_text_needed": 18,
    "notes": 40,
}


def validation_formula(options: list[str]) -> str:
    return '"' + ",".join(options) + '"'


def citation(row: pd.Series) -> str:
    pieces = [
        row.get("authors", ""),
        f"({row.get('year', '')})." if row.get("year", "") else "",
        row.get("title", ""),
        row.get("journal", ""),
    ]
    return " ".join(str(piece).strip() for piece in pieces if str(piece).strip())


def build_screening_csv() -> pd.DataFrame:
    records = read_csv_if_exists(SCORED_CSV)
    if records.empty:
        records = read_csv_if_exists(DEDUPLICATED_CSV)
    require_columns(records, ["record_id", "title", "abstract"], f"{SCORED_CSV} or {DEDUPLICATED_CSV}")

    existing_screening = read_csv_if_exists(INPUT_CSV)
    existing_by_id = existing_screening.set_index("record_id") if "record_id" in existing_screening.columns else pd.DataFrame()

    screening = pd.DataFrame()
    for column in SCREENING_COLUMNS:
        screening[column] = records[column] if column in records.columns else ""

    if not existing_by_id.empty:
        for column in PRESERVE_SCREENING_COLUMNS:
            if column in existing_by_id.columns:
                screening[column] = screening["record_id"].map(existing_by_id[column]).fillna(screening[column])

    write_csv(screening, INPUT_CSV)

    existing_full_text = read_csv_if_exists(FULL_TEXT_CSV)
    existing_full_text_by_id = existing_full_text.set_index("record_id") if "record_id" in existing_full_text.columns else pd.DataFrame()
    full_text = pd.DataFrame({"record_id": records["record_id"], "citation": records.apply(citation, axis=1)})
    for column in FULL_TEXT_COLUMNS:
        if column not in full_text.columns:
            full_text[column] = ""
    if not existing_full_text_by_id.empty:
        for column in FULL_TEXT_COLUMNS:
            if column in ["record_id", "citation"]:
                continue
            if column in existing_full_text_by_id.columns:
                full_text[column] = full_text["record_id"].map(existing_full_text_by_id[column]).fillna(full_text[column])
    write_csv(full_text[FULL_TEXT_COLUMNS], FULL_TEXT_CSV)

    return screening


def main() -> None:
    ensure_dirs()
    screening = build_screening_csv()
    require_columns(
        screening,
        [
            "record_id",
            "title",
            "abstract",
            "relevance_score",
            "human_title_abstract_decision",
            "human_title_abstract_exclusion_reason",
        ],
        str(INPUT_CSV),
    )
    if len(screening) != EXPECTED_RECORD_COUNT:
        raise ValueError(f"Expected {EXPECTED_RECORD_COUNT} screening records, found {len(screening)}.")

    screening["_sort_score"] = screening["relevance_score"].astype(str).replace("", "0").astype(float)
    screening = screening.sort_values("_sort_score", ascending=False).drop(columns=["_sort_score"])

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Screening Decisions"
    worksheet.append(list(screening.columns))
    for row in screening.itertuples(index=False, name=None):
        worksheet.append(list(row))

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E78")
    human_header_fill = PatternFill("solid", fgColor="B7791F")
    human_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(color="FFFFFF", bold=True)

    headers = [cell.value for cell in worksheet[1]]
    header_to_index = {header: index + 1 for index, header in enumerate(headers)}

    for cell in worksheet[1]:
        cell.fill = human_header_fill if cell.value in PRESERVE_SCREENING_COLUMNS else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_name, width in COLUMN_WIDTHS.items():
        if column_name in header_to_index:
            worksheet.column_dimensions[get_column_letter(header_to_index[column_name])].width = width

    for column_name in ["title", "abstract", "notes"]:
        if column_name in header_to_index:
            column_letter = get_column_letter(header_to_index[column_name])
            for cell in worksheet[f"{column_letter}2:{column_letter}{worksheet.max_row}"]:
                cell[0].alignment = Alignment(vertical="top", wrap_text=True)

    for column_name in PRESERVE_SCREENING_COLUMNS:
        if column_name in header_to_index:
            column_letter = get_column_letter(header_to_index[column_name])
            for cell in worksheet[f"{column_letter}2:{column_letter}{worksheet.max_row}"]:
                cell[0].fill = human_fill
                cell[0].alignment = Alignment(vertical="top", wrap_text=True)

    if "human_title_abstract_decision" in header_to_index:
        column_letter = get_column_letter(header_to_index["human_title_abstract_decision"])
        validation = DataValidation(type="list", formula1=validation_formula(DECISION_OPTIONS), allow_blank=True)
        validation.error = "Choose include_for_full_text, maybe, or exclude."
        validation.errorTitle = "Invalid screening decision"
        worksheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{worksheet.max_row}")

    if "human_title_abstract_exclusion_reason" in header_to_index:
        column_letter = get_column_letter(header_to_index["human_title_abstract_exclusion_reason"])
        validation = DataValidation(type="list", formula1=validation_formula(EXCLUSION_REASON_OPTIONS), allow_blank=True)
        validation.error = "Choose a listed exclusion reason, or leave blank unless the decision is exclude."
        validation.errorTitle = "Invalid exclusion reason"
        worksheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{worksheet.max_row}")

    guide = workbook.create_sheet("Screening Guide")
    guide_rows = [
        ("Decision", "Use When"),
        (
            "include_for_full_text",
            "The title/abstract suggests the record may address Medicaid/CHIP, postpartum or postnatal care, coverage/eligibility/extension/continuity/churn/redetermination, and a relevant access, continuity, utilization, behavioral health, morbidity, or equity outcome.",
        ),
        ("maybe", "The record might fit, but the title/abstract is unclear."),
        ("exclude", "The record clearly fails review scope. Every exclude must have an exclusion reason."),
        ("Screening order", "Screen highest relevance_score records first, then review the remaining records for obvious exclusions or hidden relevant papers."),
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 28
    guide.column_dimensions["B"].width = 120
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = header_font

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX} with {len(screening)} records.")


if __name__ == "__main__":
    main()
